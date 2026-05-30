"""File parser service for extracting text from various document formats."""

from abc import ABC, abstractmethod
from typing import Optional


class FileParser(ABC):
    """Abstract base class for file parsers."""

    @abstractmethod
    def supports(self, file_type: str) -> bool:
        """Check if this parser supports the given MIME type."""
        pass

    @abstractmethod
    def parse(self, file_path: str) -> str:
        """Extract text content from the given file path."""
        pass


class WordParser(FileParser):
    """Parser for .docx files using python-docx."""

    def supports(self, file_type: str) -> bool:
        return file_type in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        )

    def parse(self, file_path: str) -> str:
        try:
            from docx import Document

            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs]
            return "\n".join(paragraphs)
        except Exception as e:
            return f"[Word parse error: {e}]"


class ExcelParser(FileParser):
    """Parser for .xlsx files using openpyxl."""

    def supports(self, file_type: str) -> bool:
        return file_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        )

    def parse(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    parts.append(row_text)
            return "\n".join(parts)
        except Exception as e:
            return f"[Excel parse error: {e}]"


class PDFParser(FileParser):
    """Parser for .pdf files using pdfplumber."""

    def supports(self, file_type: str) -> bool:
        return file_type == "application/pdf"

    def parse(self, file_path: str) -> str:
        try:
            import pdfplumber

            texts = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        texts.append(page_text)
            return "\n".join(texts)
        except Exception as e:
            return f"[PDF parse error: {e}]"


class TextParser(FileParser):
    """Parser for plain text and markdown files."""

    def supports(self, file_type: str) -> bool:
        return file_type in ("text/plain", "text/markdown")

    def parse(self, file_path: str) -> str:
        try:
            with open(file_path, encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            return f"[Text parse error: {e}]"


class ImageParser(FileParser):
    """Placeholder parser for image files."""

    SUPPORTED_TYPES = (
        "image/png",
        "image/jpeg",
        "image/jpg",
        "image/gif",
        "image/webp",
        "image/bmp",
        "image/tiff",
    )

    def supports(self, file_type: str) -> bool:
        return file_type in self.SUPPORTED_TYPES

    def parse(self, file_path: str) -> str:
        return "[Image file - visual content not extracted]"


# Ordered list of parser instances
PARSERS: list[FileParser] = [
    WordParser(),
    ExcelParser(),
    PDFParser(),
    TextParser(),
    ImageParser(),
]


def get_parser(file_type: str) -> Optional[FileParser]:
    """Return the first parser that supports the given MIME type."""
    for parser in PARSERS:
        if parser.supports(file_type):
            return parser
    return None


def parse_file(file_path: str, file_type: str) -> str:
    """Convenience function to parse a file using the appropriate parser."""
    parser = get_parser(file_type)
    if parser is None:
        return f"[No parser available for file type: {file_type}]"
    return parser.parse(file_path)
